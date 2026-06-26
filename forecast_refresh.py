#!/usr/bin/env python3
"""
forecast_refresh.py — Pull all forecast data from Salesforce and Gong,
write a single JSON snapshot to reports/forecast_data.json, and persist
the previous snapshot to reports/forecast_snapshot_prev.json.

Usage:
    /Users/nick/Prospecting/.venv/bin/python3 forecast_refresh.py
"""

import base64
import json
import os
import shutil
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

sys.path.insert(0, "/Users/nick/Prospecting")
from dotenv import load_dotenv

load_dotenv("/Users/nick/Prospecting/.env")

import requests

# ── Constants ──────────────────────────────────────────────────────────────────

Q2_START = "2026-04-01"
Q2_END   = "2026-06-30"
Q2_START_DT = "2026-04-01T00:00:00Z"
Q2_END_DT   = "2026-06-30T23:59:59Z"

FY_START = "2026-01-01"   # FY2026 = calendar year 2026
FY_END   = "2026-12-31"

GONG_BASE = "https://us-64844.api.gong.io"

OUTPUT_PATH   = "/Users/nick/Prospecting/reports/forecast_data.json"
PREV_PATH     = "/Users/nick/Prospecting/reports/forecast_snapshot_prev.json"

GCS_BUCKET = os.environ.get("GCS_BUCKET", "")

QUOTAS = {
    "Andrew Miller-McKeever": 175000,
    "David Wacker":           402500,
    "Ryan Allred":             58333,
    "Ryan Reed":               58333,
    "Nick Opderbeck":              0,
    "Charlie Austin":         332500,
    "Haroon Anwar":           297500,
    "Ryan Lowe":               52500,
    "Seyar Karimi":                0,
    "Ivy Gress":                   0,
}

TEAM_QUOTA_NICK  = 694167
TEAM_QUOTA_IVY   = 682500
TEAM_TOTAL_QUOTA = 1376667

# FY annual quotas (quarterly × 4)
ANNUAL_QUOTAS = {k: v * 4 for k, v in {
    "Andrew Miller-McKeever": 175000,
    "David Wacker":           402500,
    "Ryan Allred":             58333,
    "Ryan Reed":               58333,
    "Nick Opderbeck":              0,
    "Charlie Austin":         332500,
    "Haroon Anwar":           297500,
    "Ryan Lowe":               52500,
    "Seyar Karimi":                0,
    "Ivy Gress":                   0,
}.items()}
TEAM_ANNUAL_QUOTA_NICK  = TEAM_QUOTA_NICK  * 4   # $2,776,668
TEAM_ANNUAL_QUOTA_IVY   = TEAM_QUOTA_IVY   * 4   # $2,730,000
TEAM_ANNUAL_QUOTA_TOTAL = TEAM_TOTAL_QUOTA * 4   # $5,506,668

NICK_TEAM = ["Andrew Miller-McKeever", "David Wacker", "Ryan Allred", "Ryan Reed", "Nick Opderbeck"]
IVY_TEAM  = ["Charlie Austin", "Haroon Anwar", "Ryan Lowe", "Seyar Karimi", "Ivy Gress"]
ALL_REPS  = NICK_TEAM + IVY_TEAM

# Weekly windows: auto-generated from the Sunday on/before Q2 start through today's week.
# Weeks run Sunday→Saturday so they align with Salesforce report groupings.
def _build_weekly_windows(q_start_str: str) -> list[tuple[str, str, str]]:
    q_start = date.fromisoformat(q_start_str)
    today   = date.today()
    # Sunday on or before q_start  (Mon=0 … Sun=6 → offset = (wday+1)%7)
    first_sun = q_start - timedelta(days=(q_start.weekday() + 1) % 7)
    # Sunday on or before today
    cur_sun   = today   - timedelta(days=(today.weekday()   + 1) % 7)
    windows, ws = [], first_sun
    while ws <= cur_sun:
        we    = ws + timedelta(days=6)
        label = f"{ws.month}/{ws.day}–{we.month}/{we.day}"
        windows.append((label, ws.strftime("%Y-%m-%d"), we.strftime("%Y-%m-%d")))
        ws += timedelta(days=7)
    return windows

WEEKLY_WINDOWS = _build_weekly_windows(Q2_START)
WEEK_LABELS    = [w[0] for w in WEEKLY_WINDOWS]

TARGET_WEEKLY_ACTIVITY = 206


# ── Salesforce helpers ─────────────────────────────────────────────────────────

def connect_sf():
    from simple_salesforce import Salesforce
    return Salesforce(
        username=os.environ["SF_USERNAME"],
        password=os.environ["SF_PASSWORD"],
        security_token=os.environ["SF_SECURITY_TOKEN"],
    )


def soql(sf, query: str) -> list:
    try:
        return sf.query_all(query.strip()).get("records", [])
    except Exception as e:
        print(f"  SOQL error: {e}", file=sys.stderr)
        return []


def pull_report(sf, report_id: str) -> dict:
    url  = f"{sf.base_url}analytics/reports/{report_id}"
    resp = sf._call_salesforce("GET", url)
    return resp.json()


def resolve_all_rep_ids(sf) -> dict[str, str]:
    """Return {name: sf_user_id} for all reps in ALL_REPS."""
    result = {}
    for name in ALL_REPS:
        rows = soql(sf, f"SELECT Id FROM User WHERE Name = '{name}' AND IsActive = true LIMIT 1")
        if rows:
            result[name] = rows[0]["Id"]
        else:
            print(f"  Warning: User not found in SF: {name}", file=sys.stderr)
    return result


def ids_str(id_map: dict, names: list) -> str:
    ids = [id_map[n] for n in names if n in id_map]
    return "', '".join(ids)


# ── Gong helpers ───────────────────────────────────────────────────────────────

def gong_headers() -> dict:
    key    = os.environ.get("GONG_ACCESS_KEY") or os.environ.get("GONG_API_KEY", "")
    secret = os.environ.get("GONG_ACCESS_KEY_SECRET") or os.environ.get("GONG_API_SECRET", "")
    token  = base64.b64encode(f"{key}:{secret}".encode()).decode()
    return {"Authorization": f"Basic {token}"}


# ── Section 1: forecast ────────────────────────────────────────────────────────

def build_forecast(sf, id_map: dict) -> dict:
    print("  [1/8] Pulling forecast data...")

    all_ids_str = ids_str(id_map, ALL_REPS)
    nick_ids_str = ids_str(id_map, NICK_TEAM)
    ivy_ids_str  = ids_str(id_map, IVY_TEAM)

    # Reverse id_map: sf_id -> name
    id_to_name = {v: k for k, v in id_map.items()}

    # Closed Won CQ — group by OwnerId to avoid relationship field in GROUP BY
    cw_rows = soql(sf, f"""
        SELECT OwnerId, SUM(Amount) total
        FROM Opportunity
        WHERE IsWon = true
        AND CloseDate >= {Q2_START}
        AND CloseDate <= {Q2_END}
        AND OwnerId IN ('{all_ids_str}')
        GROUP BY OwnerId
    """)
    cw_by_rep = {id_to_name.get(r["OwnerId"], r["OwnerId"]): (r.get("total") or r.get("expr0") or 0) for r in cw_rows}

    # Closed Won FY — full fiscal year (Jan 1 – Dec 31, 2026)
    fy_cw_rows = soql(sf, f"""
        SELECT OwnerId, SUM(Amount) total
        FROM Opportunity
        WHERE IsWon = true
        AND CloseDate >= {FY_START}
        AND CloseDate <= {FY_END}
        AND OwnerId IN ('{all_ids_str}')
        GROUP BY OwnerId
    """)
    fy_cw_by_rep = {id_to_name.get(r["OwnerId"], r["OwnerId"]): (r.get("total") or r.get("expr0") or 0) for r in fy_cw_rows}

    # Open opps by forecast category — group by OwnerId + ForecastCategoryName
    open_rows = soql(sf, f"""
        SELECT OwnerId, ForecastCategoryName, SUM(Amount) total
        FROM Opportunity
        WHERE IsClosed = false
        AND CloseDate >= {Q2_START}
        AND CloseDate <= {Q2_END}
        AND OwnerId IN ('{all_ids_str}')
        GROUP BY OwnerId, ForecastCategoryName
    """)

    commit_by_rep    = defaultdict(float)
    best_case_by_rep = defaultdict(float)
    pipeline_by_rep  = defaultdict(float)

    for r in open_rows:
        name = id_to_name.get(r["OwnerId"], r["OwnerId"])
        cat  = r.get("ForecastCategoryName") or ""
        amt  = r.get("total") or r.get("expr0") or 0
        if cat == "Commit":
            commit_by_rep[name] += amt
        elif cat == "Best Case":
            best_case_by_rep[name] += amt
        elif cat == "Pipeline":
            pipeline_by_rep[name] += amt

    def team_totals(team_names):
        return {
            "closed_won": sum(cw_by_rep.get(n, 0) for n in team_names),
            "commit":     sum(commit_by_rep.get(n, 0) for n in team_names),
            "best_case":  sum(best_case_by_rep.get(n, 0) for n in team_names),
            "pipeline":   sum(pipeline_by_rep.get(n, 0) for n in team_names),
        }

    by_rep = {}
    for name in ALL_REPS:
        by_rep[name] = {
            "quota":      QUOTAS.get(name, 0),
            "closed_won": cw_by_rep.get(name, 0),
            "commit":     commit_by_rep.get(name, 0),
            "best_case":  best_case_by_rep.get(name, 0),
            "pipeline":   pipeline_by_rep.get(name, 0),
        }

    nick_t = team_totals(NICK_TEAM)
    ivy_t  = team_totals(IVY_TEAM)
    comb_t = team_totals(ALL_REPS)

    print(f"    CW={comb_t['closed_won']:,.0f}  Commit={comb_t['commit']:,.0f}  BC={comb_t['best_case']:,.0f}")

    # FY by-rep enrichment
    for name in ALL_REPS:
        by_rep[name]["fy_closed_won"] = fy_cw_by_rep.get(name, 0)
        by_rep[name]["fy_quota"]      = ANNUAL_QUOTAS.get(name, 0)

    # All CQ opps (open + closed) for the deal list table
    cq_opp_rows = soql(sf, f"""
        SELECT Id, Name, Account.Name, StageName, ForecastCategoryName,
               Amount, CloseDate, OwnerId, IsClosed, IsWon
        FROM Opportunity
        WHERE CloseDate >= {Q2_START}
        AND CloseDate <= {Q2_END}
        AND OwnerId IN ('{all_ids_str}')
        ORDER BY Amount DESC NULLS LAST
    """)
    id_to_name = {v: k for k, v in id_map.items()}
    cq_deals = []
    for r in cq_opp_rows:
        acct = (r.get("Account") or {})
        cq_deals.append({
            "opp_id":    r.get("Id") or "",
            "opp_name":  r.get("Name") or "",
            "account":   acct.get("Name") or "",
            "stage":     r.get("StageName") or "",
            "category":  r.get("ForecastCategoryName") or "",
            "amount":    r.get("Amount") or 0,
            "close_date": (r.get("CloseDate") or "")[:10],
            "owner":     id_to_name.get(r.get("OwnerId") or "", ""),
            "is_won":    bool(r.get("IsWon")),
            "is_closed": bool(r.get("IsClosed")),
        })
    print(f"    CQ deal list: {len(cq_deals)} opps")

    return {
        "team_quota":  TEAM_TOTAL_QUOTA,
        "nick_team":   {"quota": TEAM_QUOTA_NICK, **nick_t},
        "ivy_team":    {"quota": TEAM_QUOTA_IVY, **ivy_t},
        "combined":    {"quota": TEAM_TOTAL_QUOTA, **comb_t},
        "by_rep":      by_rep,
        "fy_team_quotas": {
            "nick": TEAM_ANNUAL_QUOTA_NICK,
            "ivy":  TEAM_ANNUAL_QUOTA_IVY,
            "all":  TEAM_ANNUAL_QUOTA_TOTAL,
        },
        "cq_deals": cq_deals,
    }


# ── Section 2: deal_changes ────────────────────────────────────────────────────

def build_deal_changes(sf, id_map: dict) -> list:
    print("  [2/8] Pulling deal changes (last 7 days)...")

    all_ids_str = ids_str(id_map, ALL_REPS)
    since = (date.today() - timedelta(days=7)).strftime("%Y-%m-%dT00:00:00Z")

    history = soql(sf, f"""
        SELECT OpportunityId, Opportunity.Name, Opportunity.Account.Name,
               Opportunity.Owner.Name, Opportunity.Amount, Opportunity.StageName,
               Opportunity.Vol_Estimate_API_Calls_After_90_Day__c,
               Opportunity.Blended_Estimated_Cost_API_Call_CPM__c,
               Field, OldValue, NewValue, CreatedDate
        FROM OpportunityFieldHistory
        WHERE Opportunity.OwnerId IN ('{all_ids_str}')
        AND Field IN ('StageName', 'CloseDate', 'ForecastCategoryName', 'Amount')
        AND CreatedDate >= {since}
        ORDER BY CreatedDate DESC
    """)

    # Collect unique opp IDs so we can look up when each deal last changed stage
    opp_ids = list({h.get("OpportunityId") for h in history if h.get("OpportunityId")})

    # For each opp, find the most recent StageName change date (may predate the 7-day window)
    stage_entry: dict[str, str] = {}  # opp_id → ISO date of last stage change
    for i in range(0, len(opp_ids), 200):
        chunk = "', '".join(opp_ids[i:i+200])
        rows = soql(sf, f"""
            SELECT OpportunityId, CreatedDate
            FROM OpportunityFieldHistory
            WHERE OpportunityId IN ('{chunk}')
            AND Field = 'StageName'
            ORDER BY CreatedDate DESC
        """)
        for r in rows:
            oid = r.get("OpportunityId") or ""
            if oid and oid not in stage_entry:
                stage_entry[oid] = (r.get("CreatedDate") or "")[:10]

    today_str = date.today().isoformat()

    changes = []
    for h in history:
        opp    = h.get("Opportunity") or {}
        acct   = (opp.get("Account") or {}).get("Name") or ""
        owner  = (opp.get("Owner") or {}).get("Name") or ""
        oid    = h.get("OpportunityId") or ""
        entry  = stage_entry.get(oid)
        days_in = (date.fromisoformat(today_str) - date.fromisoformat(entry)).days if entry else None
        changes.append({
            "opp_id":        oid,
            "opp_name":      opp.get("Name") or "",
            "account":       acct,
            "owner":         owner,
            "field":         h.get("Field") or "",
            "old_value":     str(h.get("OldValue") or ""),
            "new_value":     str(h.get("NewValue") or ""),
            "changed_at":    (h.get("CreatedDate") or "")[:10],
            "amount":        opp.get("Amount"),
            "current_stage":   opp.get("StageName") or "",
            "days_in_stage":   days_in,
            "api_calls_month3": opp.get("Vol_Estimate_API_Calls_After_90_Day__c"),
            "blended_cpm":      opp.get("Blended_Estimated_Cost_API_Call_CPM__c"),
        })

    print(f"    {len(changes)} field history records, {len(opp_ids)} deals")
    return changes


# ── Section 3: pipeline_health ─────────────────────────────────────────────────

def build_pipeline_health(sf, id_map: dict) -> dict:
    print("  [3/8] Pulling pipeline health...")

    all_ids_str = ids_str(id_map, ALL_REPS)

    # Open pipeline CQ by rep (report) — aggs[0]=amount, aggs[1]=count
    by_rep_pipeline       = {}
    by_rep_pipeline_count = {}
    by_rep_created        = {}
    try:
        rpt = pull_report(sf, "00OVq00000DFs8fMAD")
        fact_map = rpt.get("factMap") or {}
        groupings = rpt.get("groupingsDown", {}).get("groupings", [])
        for grp in groupings:
            rep_name  = grp.get("label") or ""
            key       = f"{grp.get('key')}!T"
            cell_data = (fact_map.get(key) or {}).get("aggregates") or []
            amt   = next((float(a["value"]) for a in cell_data if isinstance(a.get("value"), float)), 0)
            count = next((int(a["value"])   for a in cell_data if isinstance(a.get("value"), int)),   0)
            by_rep_pipeline[rep_name]       = amt
            by_rep_pipeline_count[rep_name] = count
        print(f"    Open pipeline report: {len(by_rep_pipeline)} reps")
    except Exception as e:
        print(f"    Warning: open pipeline report failed: {e}", file=sys.stderr)

    # Opps created CQ (report)
    try:
        rpt2 = pull_report(sf, "00OVq00000DFeTuMAL")
        fact_map2 = rpt2.get("factMap") or {}
        groupings2 = rpt2.get("groupingsDown", {}).get("groupings", [])
        for grp in groupings2:
            rep_name = grp.get("label") or ""
            key = f"{grp.get('key')}!T"
            cell_data = (fact_map2.get(key) or {}).get("aggregates") or []
            cnt = 0
            for agg in cell_data:
                if isinstance(agg.get("value"), int):
                    cnt = agg["value"]
                    break
            if rep_name not in by_rep_pipeline:
                by_rep_pipeline[rep_name] = 0
            by_rep_created[rep_name] = cnt
        print(f"    Opps created report: {len(by_rep_created)} reps")
    except Exception as e:
        print(f"    Warning: opps created report failed: {e}", file=sys.stderr)

    # Pipeline created by week (report) — aggs[0]=amount, aggs[1]=count
    pipeline_by_week:       dict[str, dict] = {w: {} for w in WEEK_LABELS}
    pipeline_by_week_count: dict[str, dict] = {w: {} for w in WEEK_LABELS}
    try:
        rpt3 = pull_report(sf, "00OVq00000E5qabMAB")
        fact_map3  = rpt3.get("factMap") or {}
        grp_down   = rpt3.get("groupingsDown",   {}).get("groupings", [])
        grp_across = rpt3.get("groupingsAcross", {}).get("groupings", [])

        for row_grp in grp_down:
            rep_name = row_grp.get("label") or ""
            for ci, col_grp in enumerate(grp_across):
                key  = f"{row_grp.get('key')}!{col_grp.get('key')}"
                cell = (fact_map3.get(key) or {}).get("aggregates") or []
                amt   = next((float(a["value"]) for a in cell if isinstance(a.get("value"), float)), 0)
                count = next((int(a["value"])   for a in cell if isinstance(a.get("value"), int)),   0)
                if ci < len(WEEK_LABELS):
                    pipeline_by_week[WEEK_LABELS[ci]][rep_name]       = amt
                    pipeline_by_week_count[WEEK_LABELS[ci]][rep_name] = count
        print(f"    Pipeline by week report: parsed")
    except Exception as e:
        print(f"    Warning: pipeline by week report failed: {e}", file=sys.stderr)

    # Stage distribution from SOQL
    stage_dist: dict[str, dict] = {}
    try:
        stage_rows = soql(sf, f"""
            SELECT StageName, COUNT(Id) cnt, SUM(Amount) total
            FROM Opportunity
            WHERE IsClosed = false
            AND CloseDate >= {Q2_START}
            AND CloseDate <= {Q2_END}
            AND OwnerId IN ('{all_ids_str}')
            GROUP BY StageName
        """)
        for r in stage_rows:
            stage = r.get("StageName") or "Unknown"
            stage_dist[stage] = {
                "count":  r.get("cnt") or r.get("expr0") or 0,
                "amount": r.get("total") or r.get("expr1") or 0,
            }
        print(f"    Stage distribution: {len(stage_dist)} stages")
    except Exception as e:
        print(f"    Warning: stage distribution SOQL failed: {e}", file=sys.stderr)

    # Stage velocity helper — computes reached/advanced per stage for a set of opp IDs
    import re as _re
    def _stage_num(s: str) -> int:
        m = _re.match(r"^(\d+)", s.strip())
        return int(m.group(1)) if m else 0

    def _velocity_for_opps(opp_ids: list) -> dict:
        if not opp_ids:
            return {}
        all_hist = []
        for i in range(0, len(opp_ids), 200):
            chunk_str = "', '".join(opp_ids[i:i+200])
            rows = soql(sf, f"""
                SELECT OpportunityId, OldValue, NewValue
                FROM OpportunityFieldHistory
                WHERE OpportunityId IN ('{chunk_str}')
                AND Field = 'StageName'
            """)
            all_hist.extend(rows)
        reached: dict[str, set] = defaultdict(set)
        advanced: dict[str, set] = defaultdict(set)
        for h in all_hist:
            old_v  = str(h.get("OldValue") or "")
            new_v  = str(h.get("NewValue") or "")
            opp_id = h.get("OpportunityId") or ""
            if old_v: reached[old_v].add(opp_id)
            if new_v: reached[new_v].add(opp_id)
            if _stage_num(new_v) > _stage_num(old_v):
                advanced[old_v].add(opp_id)
        result = {}
        for stage in set(reached) | set(advanced):
            r, a = len(reached[stage]), len(advanced[stage])
            result[stage] = {"reached": r, "advanced": a, "cvr_pct": round(a/r*100, 1) if r else 0}
        return result

    # Stage velocity — CQ (CreatedDate >= Q2_START)
    stage_velocity:    dict[str, dict] = {}
    stage_velocity_fy: dict[str, dict] = {}
    try:
        cq_opps = soql(sf, f"""
            SELECT Id FROM Opportunity
            WHERE OwnerId IN ('{all_ids_str}')
            AND CreatedDate >= {Q2_START_DT}
        """)
        cq_ids = [r["Id"] for r in cq_opps]
        stage_velocity = _velocity_for_opps(cq_ids)
        print(f"    Stage velocity CQ: {len(stage_velocity)} stages ({len(cq_ids)} opps)")
    except Exception as e:
        print(f"    Warning: stage velocity CQ failed: {e}", file=sys.stderr)

    # Stage velocity — FY (CreatedDate >= FY_START)
    try:
        fy_opps = soql(sf, f"""
            SELECT Id FROM Opportunity
            WHERE OwnerId IN ('{all_ids_str}')
            AND CreatedDate >= {FY_START}T00:00:00Z
        """)
        fy_ids = [r["Id"] for r in fy_opps]
        stage_velocity_fy = _velocity_for_opps(fy_ids)
        print(f"    Stage velocity FY: {len(stage_velocity_fy)} stages ({len(fy_ids)} opps)")
    except Exception as e:
        print(f"    Warning: stage velocity FY failed: {e}", file=sys.stderr)

    # Build by_rep summary
    pipeline_by_rep_out = {}
    for name in ALL_REPS:
        pipeline_by_rep_out[name] = {
            "open_pipeline":       by_rep_pipeline.get(name, 0),
            "open_pipeline_count": by_rep_pipeline_count.get(name, 0),
            "opps_created_cq":     by_rep_created.get(name, 0),
        }

    return {
        "by_rep":                 pipeline_by_rep_out,
        "pipeline_by_week":       pipeline_by_week,
        "pipeline_by_week_count": pipeline_by_week_count,
        "stage_distribution":     stage_dist,
        "stage_velocity":         stage_velocity,
        "stage_velocity_fy":      stage_velocity_fy,
    }


# ── Section 4: activity ────────────────────────────────────────────────────────

def build_activity(sf, id_map: dict) -> dict:
    print("  [4/8] Pulling activity data...")

    all_ids_str = ids_str(id_map, ALL_REPS)

    # Pull all SF tasks for the full weekly window (dynamic — covers Q2 start through end of current week)
    window_start = WEEKLY_WINDOWS[0][1]   # first Sunday of Q2
    window_end   = WEEKLY_WINDOWS[-1][2]  # Saturday of current week

    tasks = soql(sf, f"""
        SELECT OwnerId, Owner.Name, ActivityDate, Type, TaskSubtype, Subject
        FROM Task
        WHERE OwnerId IN ('{all_ids_str}')
        AND ActivityDate >= {window_start}
        AND ActivityDate <= {window_end}
    """)

    def _classify(task: dict) -> str:
        t    = (task.get("Type") or "").strip()
        sub  = (task.get("TaskSubtype") or "").strip()
        subj = (task.get("Subject") or "").lower()
        t_lo = t.lower()
        # Gong auto-logs a Task for every recorded call — exclude these from the
        # outbound mix. Real call data comes from the Gong API (build_gong_calls).
        if "[gong" in subj or "gong out" in subj or "gong in" in subj:
            return None
        if "[apollo" in subj or "apollo seq" in subj:
            return "Email (Apollo)"
        if (
            "linkedin" in subj
            or "inmail" in subj
            or sub.lower() in ("linkedinmail", "linkedin")
            or "linkedin" in t_lo
            or "inmail" in t_lo
        ):
            return "LinkedIn"
        if t in ("Call", "Meeting"):
            return "Call"
        if t == "Email":
            return "Email"
        if t:
            return t
        return "Other"

    n_weeks = len(WEEKLY_WINDOWS)
    # Bucket tasks by rep and week
    by_rep_by_week: dict[str, list] = {name: [0]*n_weeks for name in ALL_REPS}
    by_type: dict[str, int] = {}
    by_type_by_rep: dict[str, dict[str, int]] = {name: {} for name in ALL_REPS}
    by_type_by_week: list[dict[str, int]] = [{} for _ in WEEKLY_WINDOWS]

    for task in tasks:
        owner_name = (task.get("Owner") or {}).get("Name") or ""
        act_date   = task.get("ActivityDate") or ""
        if not act_date or owner_name not in by_rep_by_week:
            continue
        try:
            d = date.fromisoformat(act_date[:10])
        except Exception:
            continue
        kind = _classify(task)
        if kind is None:
            continue
        by_type[kind] = by_type.get(kind, 0) + 1
        by_type_by_rep[owner_name][kind] = by_type_by_rep[owner_name].get(kind, 0) + 1
        for wi, (label, w_start, w_end) in enumerate(WEEKLY_WINDOWS):
            ws = date.fromisoformat(w_start)
            we = date.fromisoformat(w_end)
            if ws <= d <= we:
                by_rep_by_week[owner_name][wi] += 1
                by_type_by_week[wi][kind] = by_type_by_week[wi].get(kind, 0) + 1
                break

    # Log any unknown types to help debug classification gaps
    known = {"Call", "Email (Apollo)", "Email", "LinkedIn", "Other"}
    unknown_types = {k: v for k, v in by_type.items() if k not in known}
    if unknown_types:
        print(f"    Unknown activity types (may need classification): {unknown_types}", file=sys.stderr)
    print(f"    Activity types: { {k: v for k, v in sorted(by_type.items(), key=lambda x: -x[1])} }")

    # Also try the SF activity report
    try:
        rpt = pull_report(sf, "00OVq00000AHIZ3MAP")
        # Extract from report if the structure matches; otherwise fall back to SOQL data
        fact_map = rpt.get("factMap") or {}
        groupings_down   = rpt.get("groupingsDown", {}).get("groupings", [])
        groupings_across = rpt.get("groupingsAcross", {}).get("groupings", [])

        if groupings_across:
            # Matrix: rows = reps, columns = weeks
            for row_grp in groupings_down:
                rep_name = row_grp.get("label") or ""
                if rep_name not in by_rep_by_week:
                    continue
                for ci, col_grp in enumerate(groupings_across):
                    col_label = col_grp.get("label") or ""
                    key = f"{row_grp.get('key')}_{col_grp.get('key')}"
                    cell = (fact_map.get(key) or {}).get("aggregates") or []
                    cnt  = 0
                    for agg in cell:
                        if isinstance(agg.get("value"), int):
                            cnt = agg["value"]
                            break
                    # Try to match week label
                    for wi, wl in enumerate(WEEK_LABELS):
                        if wl in col_label or col_label in wl:
                            by_rep_by_week[rep_name][wi] = cnt
                            break
        print(f"    Activity report: parsed {len(groupings_down)} reps")
    except Exception as e:
        print(f"    Warning: activity report failed, using SOQL data: {e}", file=sys.stderr)

    # Opps created per week per rep
    opps_by_rep_by_week: dict[str, list] = {name: [0]*7 for name in ALL_REPS}
    id_to_name = {v: k for k, v in id_map.items()}
    try:
        for wi, (label, w_start, w_end) in enumerate(WEEKLY_WINDOWS):
            rows = soql(sf, f"""
                SELECT OwnerId, COUNT(Id) cnt
                FROM Opportunity
                WHERE OwnerId IN ('{all_ids_str}')
                AND CreatedDate >= {w_start}T00:00:00Z
                AND CreatedDate <= {w_end}T23:59:59Z
                GROUP BY OwnerId
            """)
            for r in rows:
                name = id_to_name.get(r.get("OwnerId") or "", "")
                if name in opps_by_rep_by_week:
                    opps_by_rep_by_week[name][wi] = r.get("cnt") or r.get("expr0") or 0
        print(f"    Opps by week: pulled for {len(opps_by_rep_by_week)} reps")
    except Exception as e:
        print(f"    Warning: opps by week failed: {e}", file=sys.stderr)

    return {
        "weeks":               WEEK_LABELS,
        "by_rep":              {k: v for k, v in by_rep_by_week.items()},
        "opps_by_rep_by_week": opps_by_rep_by_week,
        "target_weekly":       TARGET_WEEKLY_ACTIVITY,
        "by_type":             by_type,
        "by_type_by_rep":      by_type_by_rep,
        "by_type_by_week":     by_type_by_week,
    }


# ── Section 5: opp_sources ────────────────────────────────────────────────────

def build_opp_sources(sf, id_map: dict) -> dict:
    print("  [5/8] Pulling opp sources...")

    all_ids_str = ids_str(id_map, ALL_REPS)

    id_to_name_src = {v: k for k, v in id_map.items()}
    rows = soql(sf, f"""
        SELECT Type, OwnerId, COUNT(Id) cnt, SUM(Amount) total_amount
        FROM Opportunity
        WHERE CreatedDate >= {Q2_START_DT}
        AND OwnerId IN ('{all_ids_str}')
        GROUP BY Type, OwnerId
    """)

    by_type: dict[str, dict] = defaultdict(lambda: {"count": 0, "amount": 0})
    by_rep_type: dict[str, dict] = defaultdict(lambda: defaultdict(int))

    for r in rows:
        t    = r.get("Type") or "Unknown"
        name = id_to_name_src.get(r.get("OwnerId") or "", "")
        cnt  = r.get("cnt") or r.get("expr0") or 0
        amt  = r.get("total_amount") or r.get("expr1") or 0

        by_type[t]["count"]  += cnt
        by_type[t]["amount"] += (amt or 0)
        by_rep_type[name][t] += cnt

    print(f"    Opp sources: {len(by_type)} types")
    return {
        "by_type":     dict(by_type),
        "by_rep_type": {k: dict(v) for k, v in by_rep_type.items()},
    }


# ── Section 6: gong_calls ─────────────────────────────────────────────────────

MEETINGS_LOG_BLOB = "meetings_log.json"

NICK_TEAM_EMAILS = {
    "andrew.miller-mckeever@you.com": "Andrew Miller-McKeever",
    "david.wacker@you.com":           "David Wacker",
    "ryan.allred@you.com":            "Ryan Allred",
    "ryan.reed@you.com":              "Ryan Reed",
    "nick.opderbeck@you.com":         "Nick Opderbeck",
}
IVY_TEAM_EMAILS = {
    "charlie.austin@you.com":  "Charlie Austin",
    "haroon.anwar@you.com":    "Haroon Anwar",
    "ryan.lowe@you.com":       "Ryan Lowe",
    "seyar.karimi@you.com":    "Seyar Karimi",
    "ivy.gress@you.com":       "Ivy Gress",
}
ALL_REP_EMAILS = {**NICK_TEAM_EMAILS, **IVY_TEAM_EMAILS}
YOU_DOMAIN = "you.com"


def build_gong_calls() -> dict:
    """Returns {'calls': [...], 'meetings_14d': {'by_rep': {...}, ...}}"""
    print("  [6/8] Pulling Gong calls (last 14 days)...")

    headers = gong_headers()
    now     = datetime.now(tz=timezone.utc)
    from_dt = now - timedelta(days=14)

    from_str = from_dt.strftime("%Y-%m-%dT%H:%M:%SZ")
    to_str   = now.strftime("%Y-%m-%dT%H:%M:%SZ")

    # Step 1: pull call list
    calls_raw = []
    cursor = None
    try:
        while True:
            params: dict = {"fromDateTime": from_str, "toDateTime": to_str}
            if cursor:
                params["cursor"] = cursor
            resp = requests.get(
                f"{GONG_BASE}/v2/calls",
                headers=headers,
                params=params,
                timeout=30,
            )
            resp.raise_for_status()
            data = resp.json()
            calls_raw.extend(data.get("calls", []))
            cursor = data.get("records", {}).get("cursor")
            if not cursor:
                break
        print(f"    Gong: {len(calls_raw)} calls fetched")
    except Exception as e:
        print(f"    Warning: Gong call list failed: {e}", file=sys.stderr)
        return [{"error": str(e)}]

    # Step 2: try to enrich with extensive data (attendees)
    call_ids = [c.get("id") for c in calls_raw if c.get("id")]
    extensive_map: dict[str, dict] = {}
    if call_ids:
        try:
            for i in range(0, len(call_ids), 100):
                chunk = call_ids[i:i+100]
                resp2 = requests.post(
                    f"{GONG_BASE}/v2/calls/extensive",
                    headers={**headers, "Content-Type": "application/json"},
                    json={"filter": {"callIds": chunk}, "contentSelector": {"exposedFields": {"parties": True}}},
                    timeout=30,
                )
                if resp2.ok:
                    for c in resp2.json().get("calls", []):
                        extensive_map[c["metaData"]["id"]] = c
        except Exception as e:
            print(f"    Warning: Gong extensive call fetch failed: {e}", file=sys.stderr)

    # Step 3: parse, normalize, and count external meetings per rep
    results = []
    meetings_by_rep: dict[str, int] = {name: 0 for name in ALL_REP_EMAILS.values()}
    accounts_by_rep: dict[str, set] = {name: set() for name in ALL_REP_EMAILS.values()}

    for call in calls_raw:
        call_id    = call.get("id") or ""
        title      = call.get("title") or ""
        scheduled  = call.get("scheduled") or call.get("started") or ""
        duration   = call.get("duration") or 0
        url        = call.get("url") or ""

        # Parse account name from title "You<>AccountName"
        account = ""
        if "<>" in title:
            parts = title.split("<>", 1)
            account = parts[1].strip() if len(parts) > 1 else ""
        elif "You - " in title:
            account = title.split("You - ", 1)[1].strip()

        # Identify parties: rep attendees and whether any external party attended
        ext = extensive_map.get(call_id) or {}
        parties = ext.get("parties") or call.get("parties") or []
        rep_name = ""
        has_external = False
        for party in parties:
            email  = (party.get("emailAddress") or "").lower()
            domain = email.split("@")[1] if "@" in email else ""
            if email in ALL_REP_EMAILS:
                if not rep_name:
                    rep_name = ALL_REP_EMAILS[email]
            elif domain and domain != YOU_DOMAIN:
                has_external = True

        # Count as external meeting for the rep if at least one non-you.com party attended
        if rep_name and has_external:
            meetings_by_rep[rep_name] = meetings_by_rep.get(rep_name, 0) + 1
            if account:
                accounts_by_rep[rep_name].add(account)

        results.append({
            "title":        title,
            "account":      account,
            "date":         scheduled[:10] if scheduled else "",
            "duration_min": round(duration / 60, 1) if duration else 0,
            "url":          url,
            "rep":          rep_name,
            "has_external": has_external,
        })

    nick_total = sum(meetings_by_rep.get(n, 0) for n in NICK_TEAM_EMAILS.values())
    ivy_total  = sum(meetings_by_rep.get(n, 0) for n in IVY_TEAM_EMAILS.values())
    meetings_14d = {
        "by_rep":          meetings_by_rep,
        "accounts_by_rep": {k: sorted(v) for k, v in accounts_by_rep.items()},
        "nick_team_total": nick_total,
        "ivy_team_total":  ivy_total,
        "period_start":    from_dt.strftime("%Y-%m-%d"),
        "period_end":      now.strftime("%Y-%m-%d"),
    }
    print(f"    Gong: {len(results)} calls parsed, {nick_total + ivy_total} external meetings (14d)")
    return {"calls": results, "meetings_14d": meetings_14d}


# ── Meetings log (GCS accumulator) ───────────────────────────────────────────

def persist_meetings_log(meetings_14d: dict) -> list[dict]:
    """
    Read meetings_log.json from GCS (or start fresh), upsert this week's entry
    keyed by the Monday of the current week, and re-upload. Returns the full log.
    """
    from datetime import date, timedelta

    today     = date.today()
    week_key  = (today - timedelta(days=today.weekday())).isoformat()  # Monday ISO
    week_label = (today - timedelta(days=today.weekday())).strftime("Wk of %b %-d")

    entry = {
        "week":            week_key,
        "week_label":      week_label,
        "recorded_at":     datetime.now(tz=timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "period_start":    meetings_14d["period_start"],
        "period_end":      meetings_14d["period_end"],
        "by_rep":          meetings_14d["by_rep"],
        "accounts_by_rep": meetings_14d.get("accounts_by_rep", {}),
        "nick_team_total": meetings_14d["nick_team_total"],
        "ivy_team_total":  meetings_14d["ivy_team_total"],
    }

    log: list[dict] = []

    if GCS_BUCKET:
        try:
            from google.cloud import storage as _gcs
            client = _gcs.Client()
            bucket = client.bucket(GCS_BUCKET)
            blob   = bucket.blob(MEETINGS_LOG_BLOB)
            if blob.exists():
                existing = json.loads(blob.download_as_text())
                log = existing if isinstance(existing, list) else existing.get("entries", [])
        except Exception as e:
            print(f"    Warning: could not read meetings_log from GCS: {e}", file=sys.stderr)

    # Upsert: replace existing entry for same week, or append
    log = [e for e in log if e.get("week") != week_key]
    log.append(entry)
    log.sort(key=lambda e: e.get("week", ""))

    if GCS_BUCKET:
        try:
            from google.cloud import storage as _gcs
            client = _gcs.Client()
            bucket = client.bucket(GCS_BUCKET)
            blob   = bucket.blob(MEETINGS_LOG_BLOB)
            blob.upload_from_string(json.dumps(log, indent=2), content_type="application/json")
            print(f"    Meetings log updated: {len(log)} weeks → gs://{GCS_BUCKET}/{MEETINGS_LOG_BLOB}")
        except Exception as e:
            print(f"    Warning: could not write meetings_log to GCS: {e}", file=sys.stderr)
    else:
        # Local fallback
        log_path = OUTPUT_PATH.replace("forecast_data.json", "meetings_log.json")
        with open(log_path, "w") as f:
            json.dump(log, f, indent=2)
        print(f"    Meetings log saved locally: {log_path}")

    return log


# ── Section 7: usage_signals ──────────────────────────────────────────────────

# Accounts excluded from Usage Signals (known false positives / non-prospects)
# Add account names here (case-insensitive match against account name)
USAGE_SIGNALS_BLOCKLIST = {
    "byteplus",
    "不详",
}

def build_usage_signals(sf) -> dict:
    print("  [7/8] Pulling usage signals...")

    users = soql(sf, """
        SELECT Account__c, Account__r.Name, Account__r.Account_Tier__c,
               Email__c, API_Calls_Last_7_Days__c, API_Calls_Last_30_Days__c,
               API_Calls_per_User_All_Time__c, First_API_Call_Date__c, Last_API_Call_Date__c
        FROM Product_User__c
        WHERE Account__r.Account_Tier__c IN ('1. TARGET ACCOUNT', '2.A')
        ORDER BY API_Calls_Last_30_Days__c DESC NULLS LAST
        LIMIT 200
    """)

    # Group by account
    by_account: dict[str, dict] = {}
    for u in users:
        acct_id   = u.get("Account__c") or ""
        acct_name = (u.get("Account__r") or {}).get("Name") or ""
        tier      = (u.get("Account__r") or {}).get("Account_Tier__c") or ""
        calls_7d  = u.get("API_Calls_Last_7_Days__c") or 0
        calls_30d = u.get("API_Calls_Last_30_Days__c") or 0
        last_call = u.get("Last_API_Call_Date__c") or ""

        if acct_id not in by_account:
            by_account[acct_id] = {
                "account":      acct_name,
                "tier":         tier,
                "total_7d":     0,
                "total_30d":    0,
                "active_users_30d": 0,
                "days_dark":    None,
                "user_emails":  [],
                "last_call":    "",
            }
        by_account[acct_id]["total_7d"]  += calls_7d
        by_account[acct_id]["total_30d"] += calls_30d
        if calls_30d > 0:
            by_account[acct_id]["active_users_30d"] += 1
        if last_call:
            if not by_account[acct_id]["last_call"] or last_call > by_account[acct_id]["last_call"]:
                by_account[acct_id]["last_call"] = last_call
        email = u.get("Email__c") or ""
        if email:
            by_account[acct_id]["user_emails"].append(email)

    # Compute days_dark
    today = date.today()
    for acct_id, acc in by_account.items():
        lc = acc.get("last_call") or ""
        if lc:
            try:
                last_d = date.fromisoformat(lc[:10])
                acc["days_dark"] = (today - last_d).days
            except Exception:
                pass

    # Pull open opp counts and closed-won revenue per account
    acct_ids = list(by_account.keys())
    open_opp_map: dict[str, int] = {}
    is_customer_map: dict[str, bool] = {}
    if acct_ids:
        try:
            for i in range(0, len(acct_ids), 500):
                chunk = acct_ids[i:i+500]
                chunk_str = "', '".join(chunk)
                # Open opps
                rows = soql(sf, f"""
                    SELECT AccountId, COUNT(Id) cnt
                    FROM Opportunity
                    WHERE IsClosed = false
                    AND AccountId IN ('{chunk_str}')
                    GROUP BY AccountId
                """)
                for r in rows:
                    # simple_salesforce may return alias as 'cnt' or fall back to 'expr0'
                    cnt = r.get("cnt") or r.get("expr0") or 0
                    open_opp_map[r["AccountId"]] = cnt
                # Closed-won revenue (identifies existing customers)
                cw_rows = soql(sf, f"""
                    SELECT AccountId, SUM(Amount) total
                    FROM Opportunity
                    WHERE IsWon = true
                    AND AccountId IN ('{chunk_str}')
                    GROUP BY AccountId
                """)
                for r in cw_rows:
                    total = r.get("total") or r.get("expr0") or 0
                    if total > 0:
                        is_customer_map[r["AccountId"]] = True
        except Exception as e:
            print(f"    Warning: open opp count fetch failed: {e}", file=sys.stderr)

    # Pull account owner names
    owner_map: dict[str, str] = {}
    if acct_ids:
        try:
            for i in range(0, len(acct_ids), 500):
                chunk_str = "', '".join(acct_ids[i:i+500])
                owner_rows = soql(sf, f"""
                    SELECT Id, Owner.Name FROM Account WHERE Id IN ('{chunk_str}')
                """)
                for r in owner_rows:
                    owner_map[r["Id"]] = (r.get("Owner") or {}).get("Name") or ""
        except Exception as e:
            print(f"    Warning: account owner fetch failed: {e}", file=sys.stderr)

    # Compute signals
    all_accounts = []
    for acct_id, acc in by_account.items():
        acct_lower = acc["account"].lower()
        if any(blocked in acct_lower for blocked in USAGE_SIGNALS_BLOCKLIST):
            continue
        signals = []
        has_opp     = open_opp_map.get(acct_id, 0) > 0
        is_customer = is_customer_map.get(acct_id, False)

        if acc["total_7d"] > 0 and acc["days_dark"] is not None and acc["days_dark"] <= 7:
            signals.append("new_activity")
        if acc["total_30d"] > acc["total_7d"] * 2 and acc["total_7d"] > 0:
            signals.append("growth")
        if acc["active_users_30d"] >= 3:
            signals.append("multi_user")
        # Sales gap: meaningful usage (>100 calls) with no open opp and not an existing customer
        if acc["total_30d"] > 100 and not has_opp and not is_customer:
            signals.append("sales_gap")
        if acc["days_dark"] is not None and acc["days_dark"] > 30 and acc["total_30d"] == 0:
            signals.append("risk")
        if not signals:
            signals.append("no_signal")

        all_accounts.append({
            "account":         acc["account"],
            "account_id":      acct_id,
            "owner":           owner_map.get(acct_id, ""),
            "tier":            acc["tier"],
            "total_7d":        acc["total_7d"],
            "total_30d":       acc["total_30d"],
            "active_users_30d": acc["active_users_30d"],
            "days_dark":       acc["days_dark"],
            "signals":         signals,
            "has_open_opp":    has_opp,
            "is_customer":     is_customer,
        })

    tier1 = [a for a in all_accounts if "TARGET" in a["tier"].upper() or a["tier"] == "1. TARGET ACCOUNT"]
    tier2a = [a for a in all_accounts if "2.A" in a["tier"]]

    # Sort by total_30d descending
    all_accounts.sort(key=lambda x: x["total_30d"], reverse=True)
    tier1.sort(key=lambda x: x["total_30d"], reverse=True)
    tier2a.sort(key=lambda x: x["total_30d"], reverse=True)

    print(f"    Usage signals: {len(all_accounts)} accounts ({len(tier1)} T1, {len(tier2a)} T2A)")
    return {
        "tier1":  tier1,
        "tier2a": tier2a,
        "all":    all_accounts,
    }


# ── Section 8: paygo ──────────────────────────────────────────────────────────

PAYGO_REPORT_ID     = "00OVq00000D8YD7MAN"
PAYGO_TARGET_TEAM   = 20   # per team
PAYGO_TARGET_TOTAL  = 40   # combined

# Deals that qualify as PayGo but don't follow the standard naming convention.
# Each entry is a Salesforce Opportunity ID.
PAYGO_EXCEPTION_IDS = {
    "006Vq00000ZPkhpIAD",  # HPE Juniper Networks — Web+News Search API
    "006Vq00000aFCvOIAW",  # Reflection.ai — API Credit | Search API
}

def build_paygo(sf, id_map: dict) -> dict:
    print("  [8/9] Pulling PayGo data...")

    all_ids_str = ids_str(id_map, ALL_REPS)
    id_to_name  = {v: k for k, v in id_map.items()}

    # ── Count PayGo deals per rep from the CW report ──────────────────────────
    count_by_rep: dict[str, int] = {n: 0 for n in ALL_REPS}
    deals: list[dict] = []

    try:
        rpt = pull_report(sf, PAYGO_REPORT_ID)
        gd  = rpt.get("groupingsDown", {}).get("groupings", [])

        for rep_grp in gd:
            rep_name = rep_grp.get("label") or ""
            sub_opps = rep_grp.get("groupings", [])

            paygo_opps = [o for o in sub_opps if "PayGo" in (o.get("label") or "")]
            if rep_name in count_by_rep:
                count_by_rep[rep_name] = len(paygo_opps)

            # Parse deal label: "Account | Type | $Amount | Product | Sub-product"
            for opp in paygo_opps:
                parts = [p.strip() for p in (opp.get("label") or "").split("|")]
                account = parts[0] if len(parts) > 0 else ""
                product = parts[3] if len(parts) > 3 else "API Credit - PayGo"
                deals.append({
                    "account": account,
                    "rep":     rep_name,
                    "product": product,
                    "opp_key": opp.get("key") or "",
                })

        print(f"    PayGo report: {sum(count_by_rep.values())} deals across {sum(1 for v in count_by_rep.values() if v > 0)} reps")
    except Exception as e:
        print(f"    Warning: PayGo report failed: {e}", file=sys.stderr)

    # ── Enrich deals with tier + close date via SOQL ──────────────────────────
    try:
        exc_ids_sql = ", ".join(f"'{i}'" for i in PAYGO_EXCEPTION_IDS)
        deal_rows = soql(sf, f"""
            SELECT Id, Name, Account.Id, Account.Name, Account.Account_Tier__c,
                   OwnerId, Amount, CloseDate
            FROM Opportunity
            WHERE IsWon = true
            AND CloseDate >= {Q2_START}
            AND CloseDate <= {Q2_END}
            AND OwnerId IN ('{all_ids_str}')
            AND (Name LIKE '%PayGo%' OR Id IN ({exc_ids_sql}))
            ORDER BY CloseDate DESC
        """)
        deals = []  # rebuild from SOQL — more reliable
        for d in deal_rows:
            acct = (d.get("Account") or {})
            name_parts = [p.strip() for p in (d.get("Name") or "").split("|")]
            product = name_parts[3] if len(name_parts) > 3 else "PayGo"
            sub     = name_parts[4] if len(name_parts) > 4 else ""
            rep_name = id_to_name.get(d.get("OwnerId") or "", "")
            deals.append({
                "opp_id":     d.get("Id") or "",
                "account":    acct.get("Name") or name_parts[0] if name_parts else "",
                "account_id": acct.get("Id") or "",
                "tier":       acct.get("Account_Tier__c") or "—",
                "rep":        rep_name,
                "product":    product,
                "sub":        sub,
                "amount":     d.get("Amount") or 0,
                "close_date": (d.get("CloseDate") or "")[:10],
            })
            # Exception deals weren't in the SF report so count_by_rep missed them
            if d.get("Id") in PAYGO_EXCEPTION_IDS and rep_name in count_by_rep:
                count_by_rep[rep_name] = count_by_rep.get(rep_name, 0) + 1
    except Exception as e:
        print(f"    Warning: PayGo SOQL enrichment failed: {e}", file=sys.stderr)

    # ── All-time PAGO deals (for usage health view) ───────────────────────────
    try:
        all_pago_rows = soql(sf, f"""
            SELECT Id, Name, Account.Id, Account.Name, Account.Account_Tier__c,
                   OwnerId, Amount, CloseDate,
                   Vol_Estimate_API_Calls_After_90_Day__c,
                   Volume_Estimate_API_Monthly_Calls__c,
                   Total_Potential_Volume__c
            FROM Opportunity
            WHERE IsWon = true
            AND OwnerId IN ('{all_ids_str}')
            AND (Name LIKE '%PayGo%' OR Id IN ({exc_ids_sql}))
            ORDER BY CloseDate DESC
        """)
        all_pago_deals = []
        seen_acct_ids: set = set()
        for d in all_pago_rows:
            acct = (d.get("Account") or {})
            acct_id = acct.get("Id") or ""
            # Deduplicate by account — ORDER BY CloseDate DESC keeps most recent first
            name_parts = [p.strip() for p in (d.get("Name") or "").split("|")]
            sub = name_parts[4] if len(name_parts) > 4 else ""
            if acct_id and acct_id in seen_acct_ids:
                print(f"      Skipping duplicate PAGO deal for {acct.get('Name')} ({d.get('Name')}, {d.get('CloseDate')})")
                continue
            if acct_id:
                seen_acct_ids.add(acct_id)
            all_pago_deals.append({
                "account":        acct.get("Name") or "",
                "account_id":     acct_id,
                "tier":           acct.get("Account_Tier__c") or "—",
                "rep":            id_to_name.get(d.get("OwnerId") or "", ""),
                "sub":            sub,
                "amount":         d.get("Amount") or 0,
                "close_date":     (d.get("CloseDate") or "")[:10],
                "month3_target":  d.get("Vol_Estimate_API_Calls_After_90_Day__c"),
                "steady_state":   d.get("Volume_Estimate_API_Monthly_Calls__c"),
                "volume_upside":  d.get("Total_Potential_Volume__c"),
            })
        print(f"    All-time PAGO deals: {len(all_pago_deals)} unique accounts")
    except Exception as e:
        print(f"    Warning: All-time PAGO query failed: {e}", file=sys.stderr)
        all_pago_deals = []

    # ── Aggregate by team (CQ) ───────────────────────────────────────────────
    nick_count = sum(count_by_rep.get(n, 0) for n in NICK_TEAM)
    ivy_count  = sum(count_by_rep.get(n, 0) for n in IVY_TEAM)

    by_rep = {
        name: {"count": count_by_rep.get(name, 0)}
        for name in ALL_REPS
    }

    # ── FY counts + deals ────────────────────────────────────────────────────
    fy_count_by_rep: dict[str, int] = {n: 0 for n in ALL_REPS}
    fy_deals: list[dict] = []
    try:
        fy_rows = soql(sf, f"""
            SELECT OwnerId, COUNT(Id) cnt
            FROM Opportunity
            WHERE IsWon = true
            AND CloseDate >= {FY_START}
            AND CloseDate <= {FY_END}
            AND OwnerId IN ('{all_ids_str}')
            AND (Name LIKE '%PayGo%' OR Id IN ({exc_ids_sql}))
            GROUP BY OwnerId
        """)
        for r in fy_rows:
            name = id_to_name.get(r.get("OwnerId") or "", "")
            if name in fy_count_by_rep:
                fy_count_by_rep[name] = r.get("cnt") or r.get("expr0") or 0

        fy_deal_rows = soql(sf, f"""
            SELECT Id, Name, Account.Id, Account.Name, Account.Account_Tier__c,
                   OwnerId, Amount, CloseDate
            FROM Opportunity
            WHERE IsWon = true
            AND CloseDate >= {FY_START}
            AND CloseDate <= {FY_END}
            AND OwnerId IN ('{all_ids_str}')
            AND (Name LIKE '%PayGo%' OR Id IN ({exc_ids_sql}))
            ORDER BY CloseDate DESC
        """)
        for d in fy_deal_rows:
            acct = (d.get("Account") or {})
            name_parts = [p.strip() for p in (d.get("Name") or "").split("|")]
            product = name_parts[3] if len(name_parts) > 3 else "PayGo"
            sub     = name_parts[4] if len(name_parts) > 4 else ""
            fy_deals.append({
                "account":    acct.get("Name") or (name_parts[0] if name_parts else ""),
                "account_id": acct.get("Id") or "",
                "tier":       acct.get("Account_Tier__c") or "—",
                "rep":        id_to_name.get(d.get("OwnerId") or "", ""),
                "product":    product,
                "sub":        sub,
                "amount":     d.get("Amount") or 0,
                "close_date": (d.get("CloseDate") or "")[:10],
            })
        print(f"    FY PayGo: {len(fy_deals)} deals")
    except Exception as e:
        print(f"    Warning: FY PayGo query failed: {e}", file=sys.stderr)

    fy_nick_count = sum(fy_count_by_rep.get(n, 0) for n in NICK_TEAM)
    fy_ivy_count  = sum(fy_count_by_rep.get(n, 0) for n in IVY_TEAM)
    fy_by_rep = {name: {"count": fy_count_by_rep.get(name, 0)} for name in ALL_REPS}

    return {
        "target_per_team": PAYGO_TARGET_TEAM,
        "target_total":    PAYGO_TARGET_TOTAL,
        "nick_team":  {"count": nick_count,              "target": PAYGO_TARGET_TEAM},
        "ivy_team":   {"count": ivy_count,               "target": PAYGO_TARGET_TEAM},
        "combined":   {"count": nick_count + ivy_count,  "target": PAYGO_TARGET_TOTAL},
        "by_rep":     by_rep,
        "deals":      deals,
        # FY equivalents (same targets as CQ)
        "fy_target_per_team": PAYGO_TARGET_TEAM,
        "fy_target_total":    PAYGO_TARGET_TOTAL,
        "fy_nick_team":  {"count": fy_nick_count,                "target": PAYGO_TARGET_TEAM},
        "fy_ivy_team":   {"count": fy_ivy_count,                 "target": PAYGO_TARGET_TEAM},
        "fy_combined":   {"count": fy_nick_count + fy_ivy_count, "target": PAYGO_TARGET_TOTAL},
        "fy_by_rep":     fy_by_rep,
        "fy_deals":      fy_deals,
        "all_pago_deals": all_pago_deals,
    }


# ── Section 9: wow_delta ──────────────────────────────────────────────────────

def compute_wow_delta(new_forecast: dict) -> dict:
    """Load prev snapshot if it exists and compute deltas."""
    if not os.path.exists(PREV_PATH):
        return {}
    try:
        with open(PREV_PATH) as f:
            prev = json.load(f)

        prev_forecast = prev.get("forecast") or {}
        prev_combined = prev_forecast.get("combined") or {}
        new_combined  = new_forecast.get("combined") or {}

        return {
            "closed_won_change": new_combined.get("closed_won", 0) - prev_combined.get("closed_won", 0),
            "commit_change":     new_combined.get("commit", 0) - prev_combined.get("commit", 0),
            "pipeline_change":   new_combined.get("pipeline", 0) - prev_combined.get("pipeline", 0),
            "prev_week_of":      prev.get("week_of") or "",
        }
    except Exception as e:
        print(f"  Warning: WoW delta computation failed: {e}", file=sys.stderr)
        return {}


# ── Google Sheets / Volume helpers ────────────────────────────────────────────

_SHEETS_TOKEN_FILE = os.path.join(os.path.dirname(__file__), ".credentials", "google_token_forecast.json")
_SHEETS_SCOPES = [
    "https://www.googleapis.com/auth/documents",
    "https://www.googleapis.com/auth/drive",
]

VOLUMES_SHEET_ID  = os.environ.get("VOLUMES_SHEET_ID", "1yW9ndOa1Ab-DRh82SKLvTvJmRSRpLzm_")
VOLUMES_GID       = int(os.environ.get("VOLUMES_GID", "553080642"))
VOLUMES_SHEET_GID = 2074151923

_VOL_MONTH_KEYS = [
    "2025-01","2025-02","2025-03","2025-04","2025-05","2025-06",
    "2025-07","2025-08","2025-09","2025-10","2025-11","2025-12",
    "2026-01","2026-02","2026-03","2026-04","2026-05","2026-06",
]
_TOP2_PREFIXES = ("bytedance", "duckduckgo")
_Q1_26 = ["2026-01","2026-02","2026-03"]
_Q2_26 = ["2026-04","2026-05","2026-06"]

# Quarterly targets (millions of API calls) — update each quarter
_VOL_TARGETS_M = {"q1": 32, "q2": 99, "q3": 366, "q4": 1009}
_LOGO_TARGETS  = {
    "q1": {"net_new": 5,  "total": 23},
    "q2": {"net_new": 40, "total": 54},
    "q3": {"net_new": 53, "total": 107},
    "q4": {"net_new": 59, "total": 166},
}


def _get_sheets_creds():
    from google.auth.transport.requests import Request

    # Prefer local user OAuth token (local dev / first-run auth)
    if os.path.exists(_SHEETS_TOKEN_FILE):
        from google.oauth2.credentials import Credentials
        creds = Credentials.from_authorized_user_file(_SHEETS_TOKEN_FILE, _SHEETS_SCOPES)
        if creds.expired and creds.refresh_token:
            creds.refresh(Request())
        if creds.valid:
            return creds

    # Fall back to ADC (service account in Cloud Run)
    import google.auth
    creds, _ = google.auth.default(scopes=_SHEETS_SCOPES)
    creds.refresh(Request())
    return creds


def build_volume_data() -> dict:
    """Download API call volume xlsx from Drive and parse it with openpyxl."""
    import io
    import openpyxl
    from googleapiclient.discovery import build as _gbuild
    from googleapiclient.http import MediaIoBaseDownload

    creds = _get_sheets_creds()
    drive = _gbuild("drive", "v3", credentials=creds, cache_discovery=False)

    # Download the raw xlsx file
    request = drive.files().get_media(fileId=VOLUMES_SHEET_ID)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    buf.seek(0)

    wb = openpyxl.load_workbook(buf, data_only=True, read_only=True)

    # Find the sheet: match by gid (_id), then name keywords, then first tab
    target_sheet = None
    for ws in wb.worksheets:
        if getattr(ws, '_id', None) == VOLUMES_GID:
            target_sheet = ws
            break
    if target_sheet is None:
        for name in wb.sheetnames:
            if "API" in name or "Volume" in name or "Call" in name:
                target_sheet = wb[name]
                break
    if target_sheet is None:
        target_sheet = wb.worksheets[0]
    print(f"      Reading sheet tab: '{target_sheet.title}'")

    # Read all rows
    rows = []
    for row in target_sheet.iter_rows(values_only=True):
        rows.append(list(row))

    # Find the per-customer header row: col 1 contains "Customer" AND col 6+ has a date.
    # This avoids false-matching section headers like "API Calls By Customer" (which have
    # no date cells in the date columns).
    header_idx = None
    for i, row in enumerate(rows):
        if len(row) > 6 and row[1] is not None and "customer" in str(row[1]).lower():
            # Confirm a date header exists at col 6+
            has_date = any(
                hasattr(row[c], "month") and hasattr(row[c], "day")
                for c in range(6, min(len(row), 12))
            )
            if has_date:
                header_idx = i
                break

    if header_idx is None:
        return {"error": "Could not find Customer header row in volume spreadsheet"}

    # Build month_key → column index map from date headers in the header row.
    # Dates are stored as Python datetime objects where:
    #   .month = actual month number, .day = 2-digit year (e.g. day=25 → 2025)
    header_row = rows[header_idx]
    month_col_map: dict = {}
    for col_idx, cell in enumerate(header_row):
        if col_idx < 6 or cell is None:
            continue
        if hasattr(cell, "month") and hasattr(cell, "day"):
            mk = f"{2000 + cell.day}-{cell.month:02d}"
            if mk in _VOL_MONTH_KEYS:
                month_col_map[mk] = col_idx

    def _int(s):
        if s is None:
            return 0
        try:
            return int(float(str(s).strip().replace(",", "").replace(" ", "")))
        except (ValueError, TypeError):
            return 0

    def _str(s):
        if s is None:
            return ""
        if hasattr(s, "strftime"):
            return s.strftime("%Y-%m-%d")
        return str(s).strip()

    customers = []
    for row in rows[header_idx + 1:]:
        if not row or len(row) < 2:
            break
        name = _str(row[1]) if row[1] is not None else ""
        if not name:
            break  # end of customer data
        lname = name.lower()
        if lname.startswith("total") or lname.startswith("%") or "top 2" in lname:
            continue  # skip aggregate/summary rows
        close_date = _str(row[3]) if len(row) > 3 else ""
        monthly = {mk: 0 for mk in _VOL_MONTH_KEYS}
        for mk, col_idx in month_col_map.items():
            if col_idx < len(row):
                monthly[mk] = _int(row[col_idx])
        q1_total = sum(monthly.get(m, 0) for m in _Q1_26)
        q2_total = sum(monthly.get(m, 0) for m in _Q2_26)
        customers.append({
            "name":          name,
            "close_date":    close_date,
            "monthly_calls": monthly,
            "q1_total":      q1_total,
            "q2_total":      q2_total,
            "is_top2":       any(name.lower().startswith(p) for p in _TOP2_PREFIXES),
        })

    wb.close()

    excl = [c for c in customers if not c["is_top2"]]
    monthly_excl = {mk: sum(c["monthly_calls"].get(mk, 0) for c in excl) for mk in _VOL_MONTH_KEYS}
    q2_excl = sum(monthly_excl.get(m, 0) for m in _Q2_26)
    q1_excl = sum(monthly_excl.get(m, 0) for m in _Q1_26)

    return {
        "targets_m":           _VOL_TARGETS_M,
        "logo_targets":        _LOGO_TARGETS,
        "monthly_excl_top2":   monthly_excl,
        "q2_excl_top2":        q2_excl,
        "q1_excl_top2":        q1_excl,
        "customers_excl_top2": excl,
        "top2":                sorted(set(c["name"] for c in customers if c["is_top2"])),
        "as_of":               datetime.now().strftime("%Y-%m-%d"),
    }


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    today      = date.today()
    # Next Sunday (or today if Sunday) as the "week of" label
    days_until_sunday = (6 - today.weekday()) % 7
    week_date  = today + timedelta(days=days_until_sunday)
    week_of    = week_date.strftime("%B %-d, %Y")
    generated  = datetime.now(tz=timezone.utc).isoformat()

    print(f"forecast_refresh.py — {today.strftime('%B %-d, %Y')}")
    print(f"Week of: {week_of}")
    print()

    # Connect to Salesforce
    print("Connecting to Salesforce...")
    try:
        sf = connect_sf()
        print("  Connected.")
    except Exception as e:
        print(f"FATAL: Salesforce connection failed: {e}", file=sys.stderr)
        sys.exit(1)

    print("Resolving rep IDs...")
    id_map = resolve_all_rep_ids(sf)
    print(f"  {len(id_map)} reps resolved")
    print()

    # Pull each section with individual error handling
    output: dict = {
        "generated_at": generated,
        "week_of":      week_of,
    }

    # 1. Forecast
    try:
        output["forecast"] = build_forecast(sf, id_map)
        print("  [1/8] Done: forecast")
    except Exception as e:
        print(f"  [1/8] ERROR: forecast: {e}", file=sys.stderr)
        output["forecast"] = {"error": str(e)}

    # 2. Deal changes
    try:
        output["deal_changes"] = build_deal_changes(sf, id_map)
        print(f"  [2/8] Done: deal_changes ({len(output['deal_changes'])} records)")
    except Exception as e:
        print(f"  [2/8] ERROR: deal_changes: {e}", file=sys.stderr)
        output["deal_changes"] = {"error": str(e)}

    # 3. Pipeline health
    try:
        output["pipeline_health"] = build_pipeline_health(sf, id_map)
        print("  [3/8] Done: pipeline_health")
    except Exception as e:
        print(f"  [3/8] ERROR: pipeline_health: {e}", file=sys.stderr)
        output["pipeline_health"] = {"error": str(e)}

    # 4. Activity
    try:
        output["activity"] = build_activity(sf, id_map)
        print("  [4/8] Done: activity")
    except Exception as e:
        print(f"  [4/8] ERROR: activity: {e}", file=sys.stderr)
        output["activity"] = {"error": str(e)}

    # 5. Opp sources
    try:
        output["opp_sources"] = build_opp_sources(sf, id_map)
        print("  [5/8] Done: opp_sources")
    except Exception as e:
        print(f"  [5/8] ERROR: opp_sources: {e}", file=sys.stderr)
        output["opp_sources"] = {"error": str(e)}

    # 6. Gong calls + external meetings log
    try:
        gong_result = build_gong_calls()
        output["gong_calls"]    = gong_result["calls"]
        output["meetings_14d"]  = gong_result["meetings_14d"]
        output["meetings_log"]  = persist_meetings_log(gong_result["meetings_14d"])
        n_calls    = len(output["gong_calls"])
        n_meetings = gong_result["meetings_14d"]["nick_team_total"] + gong_result["meetings_14d"]["ivy_team_total"]
        print(f"  [6/9] Done: gong_calls ({n_calls} calls, {n_meetings} external meetings tracked)")
    except Exception as e:
        print(f"  [6/9] ERROR: gong_calls: {e}", file=sys.stderr)
        output["gong_calls"]   = {"error": str(e)}
        output["meetings_14d"] = {}
        output["meetings_log"] = []

    # 7. Usage signals
    try:
        output["usage_signals"] = build_usage_signals(sf)
        print("  [7/9] Done: usage_signals")
    except Exception as e:
        print(f"  [7/9] ERROR: usage_signals: {e}", file=sys.stderr)
        output["usage_signals"] = {"error": str(e)}

    # 8. PayGo
    try:
        output["paygo"] = build_paygo(sf, id_map)
        pg = output["paygo"]
        print(f"  [8/9] Done: paygo ({pg['combined']['count']} deals, {pg['nick_team']['count']} Nick / {pg['ivy_team']['count']} Ivy)")
    except Exception as e:
        print(f"  [8/9] ERROR: paygo: {e}", file=sys.stderr)
        output["paygo"] = {"error": str(e)}

    # 9. WoW delta (reads prev snapshot; must run before we overwrite)
    try:
        output["wow_delta"] = compute_wow_delta(output.get("forecast") or {})
        print("  [9/10] Done: wow_delta")
    except Exception as e:
        print(f"  [9/10] ERROR: wow_delta: {e}", file=sys.stderr)
        output["wow_delta"] = {"error": str(e)}

    # 10. Volume data (Google Sheets)
    try:
        output["volume_data"] = build_volume_data()
        vd = output["volume_data"]
        cust_count = len(vd.get("customers_excl_top2") or [])
        print(f"  [10/10] Done: volume_data ({cust_count} customers, Q2 excl top2: {vd.get('q2_excl_top2',0):,})")
    except Exception as e:
        print(f"  [10/10] ERROR: volume_data: {e}", file=sys.stderr)
        output["volume_data"] = {"error": str(e)}

    # Persist snapshots
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

    # Copy current → prev before overwriting
    if os.path.exists(OUTPUT_PATH):
        shutil.copy2(OUTPUT_PATH, PREV_PATH)
        print(f"\nSnapshot saved: {PREV_PATH}")

    # Write new output locally
    with open(OUTPUT_PATH, "w") as f:
        json.dump(output, f, indent=2, default=str)
    print(f"Output written: {OUTPUT_PATH}")

    # Upload to GCS if configured
    if GCS_BUCKET:
        try:
            from google.cloud import storage as _gcs
            client = _gcs.Client()
            bucket = client.bucket(GCS_BUCKET)
            # Upload current
            blob = bucket.blob("forecast_data.json")
            blob.upload_from_filename(OUTPUT_PATH, content_type="application/json")
            print(f"Uploaded to gs://{GCS_BUCKET}/forecast_data.json")
            # Upload prev snapshot
            if os.path.exists(PREV_PATH):
                prev_blob = bucket.blob("forecast_snapshot_prev.json")
                prev_blob.upload_from_filename(PREV_PATH, content_type="application/json")
                print(f"Uploaded to gs://{GCS_BUCKET}/forecast_snapshot_prev.json")
        except Exception as gcs_err:
            print(f"⚠ GCS upload failed (local file still written): {gcs_err}", file=sys.stderr)

    print(f"Done. Generated at {generated}")


if __name__ == "__main__":
    main()
